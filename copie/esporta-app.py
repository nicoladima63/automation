import argparse
from dbfread import DBF
from datetime import datetime, date, timedelta, time
from config import (
    PATH_APPUNTAMENTI_DBF,
    PATH_ANAGRAFICA_DBF,
    COL_APPUNTAMENTI_DATA,
    COL_APPUNTAMENTI_ORA,
    COL_APPUNTAMENTI_ORA_FINE,
    COL_APPUNTAMENTI_IDPAZIENTE,
    COL_APPUNTAMENTI_TIPO,
    COL_APPUNTAMENTI_MEDICO,
    COL_APPUNTAMENTI_STUDIO,  # Assumo che aggiungerai questa costante con il nome 'DB_APSTUDI'
    COL_PAZIENTI_ID,
    COL_PAZIENTI_NOME,
    TIPI_APPUNTAMENTO,
    COLORI_APPUNTAMENTO,
    MEDICI,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def main():
    parser = argparse.ArgumentParser(description="Esporta appuntamenti DBF in Excel con due colonne per giorno (due studi)")
    parser.add_argument("--test-mese", type=int, help="Mese da esportare (1-12), default mese corrente")
    parser.add_argument("--esporta-csv", action="store_true", help="Flag per esportare in Excel")
    args = parser.parse_args()

    today = date.today()
    mese = args.test_mese if args.test_mese and 1 <= args.test_mese <= 12 else today.month
    anno = today.year

    start_date = date(anno, mese, 1)
    if mese == 12:
        end_date = date(anno + 1, 1, 1)
    else:
        end_date = date(anno, mese + 1, 1)

    # Carica pazienti in dizionario {id: nome}
    pazienti_dbf = DBF(PATH_ANAGRAFICA_DBF, encoding='latin-1')
    PAZIENTI_NOMI = {
        str(record[COL_PAZIENTI_ID]).strip(): record[COL_PAZIENTI_NOME].strip()
        for record in pazienti_dbf
    }

    # Carica appuntamenti in struttura: appuntamenti_per_giorno[giorno][orario][studio] = (testo, tipo_codice)
    appuntamenti_per_giorno = {}

    dbf = DBF(PATH_APPUNTAMENTI_DBF, encoding='latin-1')
    for record in dbf:
        data = record.get(COL_APPUNTAMENTI_DATA)
        if not isinstance(data, date):
            continue
        if not (start_date <= data < end_date):
            continue

        ora_inizio = record.get(COL_APPUNTAMENTI_ORA, 0)
        orario_minuti = (ora_inizio // 100) * 60 + (ora_inizio % 100)
        # Convertiamo l'orario in minuti da mezzanotte per uniformare

        cod_paziente = str(record.get(COL_APPUNTAMENTI_IDPAZIENTE, '')).strip()
        nome_paziente = PAZIENTI_NOMI.get(cod_paziente, f"ID:{cod_paziente}")

        tipo_codice = (record.get(COL_APPUNTAMENTI_TIPO) or '').strip().upper()
        tipo_descrizione = TIPI_APPUNTAMENTO.get(tipo_codice, "Nota giornaliera")

        # Prendo il valore studio (1 o 2)
        studio = record.get('DB_APSTUDI', None)  # Se il nome della colonna è diverso, aggiornalo

        if studio not in (1, 2):
            # Se manca o non è 1/2, salto o metto in studio 1 di default
            studio = 1

        giorno = data.strftime("%A %d %B %Y")  # es. "Thursday 12 June 2025"

        if giorno not in appuntamenti_per_giorno:
            appuntamenti_per_giorno[giorno] = {}

        if orario_minuti not in appuntamenti_per_giorno[giorno]:
            appuntamenti_per_giorno[giorno][orario_minuti] = {1: None, 2: None}

        ora_int = int(ora_inizio)
        testo = f"{ora_int//100:02d}:{ora_int%100:02d} {nome_paziente}"
        appuntamenti_per_giorno[giorno][orario_minuti][studio] = (testo, tipo_codice)

    if not args.esporta_csv:
        # Se non voglio esportare, stampo a video solo
        for giorno, orari in sorted(appuntamenti_per_giorno.items()):
            print(f"\n{giorno}")
            for orario_minuti in sorted(orari):
                vals = orari[orario_minuti]
                line = []
                for studio in (1, 2):
                    if vals[studio]:
                        line.append(f"{vals[studio][0]}")
                    else:
                        line.append("")
                print(f"{orario_minuti//60:02d}:{orario_minuti%60:02d} | {line[0]} | {line[1]}")
        return

    # Se esportiamo in Excel:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Appuntamenti_{mese}_{anno}"

    # Prepara lista giorni ordinata
    giorni = sorted(appuntamenti_per_giorno.keys())

    # Scrivi intestazione giorni, due colonne per giorno
    for giorno_idx, giorno in enumerate(giorni):
        base_col = 2 + giorno_idx * 2
        ws.merge_cells(start_row=1, start_column=base_col, end_row=1, end_column=base_col + 1)
        ws.cell(row=1, column=base_col).value = giorno

        # Intestazioni studi (senza scritte per non sovraccaricare)
        ws.cell(row=2, column=base_col).value = "Studio 1"
        ws.cell(row=2, column=base_col + 1).value = "Studio 2"

    # Prepara lista orari ordinata (tutti gli orari trovati)
    orari = sorted({orario for giorno in giorni for orario in appuntamenti_per_giorno[giorno]})

    # Scrivi orari nella colonna A a partire dalla riga 3
    for idx, orario_minuti in enumerate(orari, start=3):
        ore = orario_minuti // 60
        minuti = orario_minuti % 60
        ws.cell(row=idx, column=1).value = f"{int(ore):02d}:{int(minuti):02d}"


    # Scrivi i dati appuntamenti per ogni giorno e studio
    for giorno_idx, giorno in enumerate(giorni):
        base_col = 2 + giorno_idx * 2
        for row_idx, orario_minuti in enumerate(orari, start=3):
            for studio in (1, 2):
                cell = ws.cell(row=row_idx, column=base_col + (studio - 1))
                val = appuntamenti_per_giorno[giorno].get(orario_minuti, {}).get(studio)
                if val:
                    testo, tipo_codice = val
                    cell.value = testo
                    colore = COLORI_APPUNTAMENTO.get(tipo_codice, "#FFFFFF")
                    cell.fill = PatternFill(start_color=colore.lstrip('#'), end_color=colore.lstrip('#'), fill_type="solid")
                else:
                    cell.value = ""

    nome_file = f"appuntamenti_{anno}_{mese:02d}.xlsx"
    wb.save(nome_file)
    print(f"File Excel salvato come {nome_file}")

if __name__ == "__main__":
    main()
